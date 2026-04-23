import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from models.ExcelProcessor import ExcelProcessor


class TestExcelProcessor(unittest.TestCase):
    def test_related_moha_smells_with_deduplication(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / "Mapeo.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=4, value="CC")
            ws.cell(row=1, column=5, value="python:S101")
            ws.cell(row=2, column=4, value="CC")
            ws.cell(row=2, column=5, value="python:S101")
            wb.save(excel_path)

            sonar_smells = [
                {
                    "rule": "python:S101",
                    "line": 12,
                    "severity": "MAJOR",
                    "component": "src/a.py",
                    "issue_key": "k1",
                    "project": "owner_repo",
                    "textRange": {"startLine": 12, "endLine": 12},
                },
                {
                    "rule": "python:S101",
                    "line": 12,
                    "severity": "MAJOR",
                    "component": "src/a.py",
                    "issue_key": "k1",
                    "project": "owner_repo",
                    "textRange": {"startLine": 12, "endLine": 12},
                },
            ]

            processor = ExcelProcessor(str(excel_path), sonar_smells)
            related = processor.GetRelatedMohaSmell()

        self.assertEqual(len(related), 1)
        self.assertEqual(related[0]["moha_smell"], "CC")
        self.assertEqual(related[0]["source"], "sonar")

    def test_related_moha_smells_returns_empty_when_no_rule_matches(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / "Mapeo.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=4, value="CC")
            ws.cell(row=1, column=5, value="python:S101")
            wb.save(excel_path)

            sonar_smells = [
                {
                    "rule": "python:S999",
                    "line": 12,
                    "severity": "MAJOR",
                    "component": "src/a.py",
                    "issue_key": "k1",
                    "project": "owner_repo",
                    "textRange": {"startLine": 12, "endLine": 12},
                }
            ]

            processor = ExcelProcessor(str(excel_path), sonar_smells)
            related = processor.GetRelatedMohaSmell()

        self.assertEqual(related, [])


if __name__ == "__main__":
    unittest.main()
