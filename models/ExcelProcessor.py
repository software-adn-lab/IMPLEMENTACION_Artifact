from openpyxl import load_workbook
class ExcelProcessor:
    def __init__(self, file_path, listaSonarSmells):
        self.file_path = file_path
        self.listaSonarSmells = listaSonarSmells

    def GetRelatedMohaSmell(self):
        relatedSmells = []
        vistos = set() # Inicializamos el control de duplicados.
        
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        for fila in range(2, sheet.max_row + 1):

            sonar_rule = sheet.cell(row=fila, column=5).value
            moha_smell = sheet.cell(row=fila, column=4).value

            for smell in self.listaSonarSmells:
                # Extrae la regla del Excel.
                smell_rule = smell['rule']
                linea = smell.get('line')
                componente = smell.get('component')
                
                if smell_rule == sonar_rule:
                    # Creamos la llave con la combinacion exacta.
                    identificador = (moha_smell, sonar_rule, linea, componente)
                    
                    # Solo agregamos si esta combinacion NO esta en vistos.
                    if identificador not in vistos:
                        relatedSmells.append({
                            'moha_smell': moha_smell,
                            'sonar_rule': sonar_rule,
                            'line': linea,
                            'severity': smell.get('severity'),
                            'component': componente,
                            'issue_key': smell.get('issue_key'),
                            'project': smell.get('project'),
                            'textRange': smell.get('textRange')
                        })
                        
                        # Registramos la combinacion para no volver a agregarla.
                        vistos.add(identificador) 
                        
                        

        return relatedSmells

            
        